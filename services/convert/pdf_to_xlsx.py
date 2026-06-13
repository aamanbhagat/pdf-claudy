"""PDF -> Excel: pull tables out of each page into a sheet.

Real table extraction with pdfplumber; when a page has no detectable table we
fall back to dumping its text lines into column A so nothing is lost.
"""
import sys

import openpyxl
import pdfplumber

inp, out = sys.argv[1], sys.argv[2]
wb = openpyxl.Workbook()
wb.remove(wb.active)  # start clean; add a sheet per page

with pdfplumber.open(inp) as pdf:
    for i, page in enumerate(pdf.pages, 1):
        ws = wb.create_sheet(title=f"Page {i}"[:31])
        row = 1
        tables = page.extract_tables() or []
        if tables:
            for table in tables:
                for cells in table:
                    for col, val in enumerate(cells, 1):
                        if val is not None:
                            ws.cell(row=row, column=col, value=str(val).strip())
                    row += 1
                row += 1  # blank line between tables
        else:
            for line in (page.extract_text() or "").splitlines():
                ws.cell(row=row, column=1, value=line)
                row += 1

if not wb.sheetnames:
    wb.create_sheet(title="Sheet1")  # openpyxl requires at least one sheet
wb.save(out)
